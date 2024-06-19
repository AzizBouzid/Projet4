from views.tour_views import ChoixJoueurView
from tinydb import TinyDB, Query
from openpyxl import Workbook
from openpyxl.styles import Font
from operator import itemgetter

Valider = Query()
BDD = TinyDB('data/tournois.json', indent=4)
tournoi = BDD.table('Tournois')
tour = BDD.table('Tour')
BD = TinyDB('data/joueurs.json', indent=4)
joueur = BD.table('Joueurs')

CYAN = '\033[96m'
BOLD = '\033[1m'
END = '\033[0m'
GREEN = '\033[92m'
BLUE = '\033[94m'


class RapportJoueur():
    '''Création du rapport de tous les joueurs par ordre alphabétique
       et l'exporter'''
    def rapport_joueur_terminal(self):
        print(CYAN + BOLD + '\n\tListe des joueurs: ' + END + '\n')
        table_joueur = joueur.all()
        result = sorted(table_joueur, key=itemgetter('Nom', 'Prenom'))
        for index, item in enumerate(result):
            table_joueur[index] = (
                str(item).replace("{", "").replace("}", "").replace(
                    "'", "").replace(",", "\n")
            )
            print(GREEN + BOLD + "------ Joueur", index+1,
                  "------" '\n' + END, table_joueur[index] + '\n')
        RapportJoueur.choix(self)

    def choix(self):
        self.reponse = ChoixJoueurView().export()
        reponse = self.reponse
        if reponse:
            RapportJoueur.rapport_joueur_csv(self)

    def rapport_joueur_csv(self):
        table_joueurs = joueur.all()
        table_joueurs = sorted(table_joueurs, key=itemgetter(
            'Nom', 'Prenom'), reverse=True)
        liste_csv = []
        liste = []
        for row in table_joueurs:
            liste_csv = []
            for cle, valeur in row.items():
                liste_csv.append(valeur)
            liste.insert(0, liste_csv)
        Ajout = ['Identifiant', 'Nom', 'Prénom', 'date de naissance',
                 'Adresse', 'Mail', 'Telephone', 'Statut', 'Score']
        liste.insert(0, Ajout)
        wb = Workbook()
        ws = wb.active
        for row in liste:
            ws.append(row)
        ft = Font(bold=True)
        for row in ws["A1:I1"]:
            for cell in row:
                cell.font = ft
        wb.save("rapports/Joueurs.xlsx")

    def run(self):
        self.rapport_joueur_terminal()


class RapportTournoi():
    '''Création du rapport de tous les tournois et l'exporter'''
    def rapport_tournoi_terminal(self):
        print(CYAN + BOLD + '\n\tListe des tournois: ' + END + '\n')
        table_tournoi = tournoi.all()
        items = list(table_tournoi[0].items())
        del items[7]
        table_tournoi = [dict(items)]
        result = sorted(table_tournoi, key=itemgetter('Nom'))
        for index, item in enumerate(result):
            table_tournoi[index] = (
                str(item).replace("{", "").replace("}", "").replace(
                    "'", "").replace(",", "\n")
            )
            print(GREEN + BOLD + "------ Tournois", index+1,
                  "------" '\n' + END, table_tournoi[index] + '\n')
        RapportTournoi.choix(self)

    def choix(self):
        self.reponse = ChoixJoueurView().export()
        reponse = self.reponse
        if reponse:
            RapportTournoi.rapport_tournoi_csv(self)

    def rapport_tournoi_csv(self):
        table_tournoi = tournoi.all()
        table_tournoi = sorted(table_tournoi, key=itemgetter('Nom'))
        liste_csv = []
        liste = []
        for row in table_tournoi:
            liste_csv = []
            for cle, valeur in row.items():
                liste_csv.append(valeur)
            del liste_csv[7]
            liste.insert(0, liste_csv)
        Ajout = ['Nom', 'Lieu', 'Date de debut',
                 'Date de fin', 'Nombre de tour', 'Tour actuel', 'Description']
        liste.insert(0, Ajout)
        wb = Workbook()
        ws = wb.active
        for row in liste:
            ws.append(row)
        ft = Font(bold=True)
        for row in ws["A1:G1"]:
            for cell in row:
                cell.font = ft
        wb.save("rapports/Tournois.xlsx")

    def run(self):
        self.rapport_tournoi_terminal()


class RapportNomDateTournoil():
    '''Création du rapport nom et dates d’un tournoi donné et l'exporter'''
    def __init__(self):
        self.nom_tournoi = ChoixJoueurView().ajout_nom_tournoi()

    def nom_date_tournoi_terminal(self):
        nom_tournoi = self.nom_tournoi
        print(CYAN + BOLD + '\n\tNom date du tournoi: ' + END + '\n')
        print('Nom du tournoi: ', nom_tournoi)
        date_d = tournoi.get(Valider.Nom == nom_tournoi).get(
            'Date_debut')
        print("Date_debut:     ", date_d)
        date_f = tournoi.get(Valider.Nom == nom_tournoi).get(
            'Date_debut')
        print("Date_fin:       ", date_f + '\n')
        RapportNomDateTournoil.choix(self)

    def choix(self):
        self.reponse = ChoixJoueurView().export()
        reponse = self.reponse
        if reponse:
            RapportNomDateTournoil.nom_date_tournoi_csv(self)

    def nom_date_tournoi_csv(self):
        nom_tournoi = self.nom_tournoi
        liste_csv = []
        liste = []
        date_f = tournoi.get(Valider.Nom == nom_tournoi).get(
            'Date_fin')
        liste_csv.insert(0, date_f)
        date_d = tournoi.get(Valider.Nom == nom_tournoi).get(
            'Date_debut')
        liste_csv.insert(0, date_d)
        liste_csv.insert(0, nom_tournoi)
        liste.insert(0, liste_csv)
        Ajout = ['Nom Tournoi', 'Date debut', 'Date fin']
        liste.insert(0, Ajout)
        wb = Workbook()
        ws = wb.active
        for row in liste:
            ws.append(row)
        ft = Font(bold=True)
        for row in ws["A1:C1"]:
            for cell in row:
                cell.font = ft
        wb.save("rapports/Nom_Date_Tournoi_%s.xlsx" % nom_tournoi.capitalize())

    def run(self):
        self.nom_date_tournoi_terminal()


class RapportJoueurTournoi():
    '''Création du rapport des joueurs du tournoi par ordre alphabétique
       et l'exporter'''
    def __init__(self):
        self.nom_tournoi = ChoixJoueurView().ajout_nom_tournoi()

    def joueur_tournoi_terminal(self):
        nom_tournoi = self.nom_tournoi
        table_tournoi = tournoi.get(Valider.Nom == nom_tournoi).get(
            'Identifiant_national')
        liste_csv = []
        for id in table_tournoi:
            liste_joueur = []
            nom = joueur.get(Valider.Identifiant_national == id[0]).get('Nom')
            liste_joueur.append(nom)
            prenom = joueur.get(Valider.Identifiant_national ==
                                id[0]).get('Prenom')
            liste_joueur.append(prenom)
            liste_csv.insert(0, liste_joueur)
            liste_trier = sorted(
                liste_csv)
        print(CYAN + BOLD + "\n\tListe des joueurs dans un tournoi :\n" + END)
        for g in liste_trier:
            print("Nom:       ", g[0])
            print("Prénom:    ", g[1])
            print(GREEN + BOLD + "--------------------" + END)
        RapportJoueurTournoi.choix(self)

    def choix(self):
        self.reponse = ChoixJoueurView().export()
        reponse = self.reponse
        if reponse:
            RapportJoueurTournoi.joueur_tournoi_csv(self)

    def joueur_tournoi_csv(self):
        nom_tournoi = self.nom_tournoi
        identifiant = tournoi.get(Valider.Nom == nom_tournoi).get(
            'Identifiant_national')
        liste_csv = []
        for id in identifiant:
            liste_joueur = []
            nom = joueur.get(Valider.Identifiant_national == id[0]).get('Nom')
            liste_joueur.append(nom)
            prenom = joueur.get(Valider.Identifiant_national ==
                                id[0]).get('Prenom')
            liste_joueur.append(prenom)
            liste_csv.insert(0, liste_joueur)
            liste_trier = sorted(
                liste_csv)
        Ajout = ['Nom', 'Prénom']
        liste_trier.insert(0, Ajout)
        wb = Workbook()
        ws = wb.active
        for row in liste_trier:
            ws.append(row)
        ft = Font(bold=True)
        for row in ws["A1:B1"]:
            for cell in row:
                cell.font = ft
        wb.save("rapports/Joueurs_Tournoi_%s.xlsx" % nom_tournoi.capitalize())

    def run(self):
        self.joueur_tournoi_terminal()


class RapportTourTournoi():
    '''Création du rapport de tous les tours du tournoi et de tous les matchs
       du tour et l'exporter'''
    def __init__(self):
        self.nom_tournoi = ChoixJoueurView().ajout_nom_tournoi()

    def tour_tournoi_terminal(self):
        nom_tournoi = self.nom_tournoi
        table_tour = tour.all()
        champ = ""
        print(BLUE + BOLD + "\n\tListe des tours dans un tournoi :\n" + END)
        for info_tour in table_tour:
            champ += GREEN + BOLD + f"Tour: {info_tour['Tour']}\n" + END
            champ += f"Tournoi: {nom_tournoi}\n"
            champ += f"Date debut: {info_tour['Date_debut']}\n"
            champ += f"Heure debut: {info_tour['Heure_debut']}\n"
            champ += f"Date fin: {info_tour['Date_fin']}\n"
            champ += f"Heure debut: {info_tour['Heure_fin']}\n"
            champ += f"{"List des matchs: "}\n"
            i = 0
            x = 0
            champ += CYAN + BOLD + f'\tMatch:  {x+1}\n' + END
            while i < 2:
                champ += f'\t\tNom:  {info_tour['List_match'][0][0][i][1]}\
                'f'\tPrénom:  {info_tour['List_match'][0][0][i][2]}\
                'f'\t  Score:  {info_tour['List_match'][0][0][i][3]}\n'
                i += 1
            champ += CYAN + BOLD + f'\tMatch:  {x+2}\n' + END
            while x < 2:
                champ += f'\t\tNom:  {info_tour['List_match'][0][1][x][1]}\
                'f'\tPrénom:  {info_tour['List_match'][0][1][x][2]}\
                'f'\t  Score:  {info_tour['List_match'][0][1][x][3]}\n'
                x += 1
        print(champ)
        RapportTourTournoi.choix(self)

    def choix(self):
        self.reponse = ChoixJoueurView().export()
        reponse = self.reponse
        if reponse:
            RapportTourTournoi.tour_tournoi_txt(self)

    def tour_tournoi_txt(self):
        nom_tournoi = self.nom_tournoi
        table_tour = tour.all()
        champ = ""
        print("\n\tListe des tours dans un tournoi :\n")
        for info_tour in table_tour:
            champ += f"Tour: {info_tour['Tour']}\n"
            champ += f"Tournoi: {nom_tournoi}\n"
            champ += f"Date debut: {info_tour['Date_debut']}\n"
            champ += f"Heure debut: {info_tour['Heure_debut']}\n"
            champ += f"Date fin: {info_tour['Date_fin']}\n"
            champ += f"Heure debut: {info_tour['Heure_fin']}\n"
            champ += f"{"List des matchs: "}\n"
            i = 0
            x = 0
            champ += f'\tMatch: {x+1}\n'
            while i < 2:
                champ += f'\t\tNom: {info_tour['List_match'][0][0][i][1]}\
                'f'\tPrénom: {info_tour['List_match'][0][0][i][2]}\
                'f'\tScore: {info_tour['List_match'][0][0][i][3]}\n'
                i += 1
            champ += f'\tMatch: {x+2}\n'
            while x < 2:
                champ += f'\t\tNom: {info_tour['List_match'][0][1][x][1]}\
                'f'\tPrénom: {info_tour['List_match'][0][1][x][2]}\
                'f'\tScore: {info_tour['List_match'][0][1][x][3]}\n'
                x += 1
        try:
            with open(
                f"{"rapports/Tours_Tournoi_%s" %
                    nom_tournoi.capitalize()}.txt",
                "w", encoding="utf-8"
            ) as file:
                file.write(champ)
        except Exception:
            return False
        else:
            return True

    def run(self):
        self.tour_tournoi_terminal()
